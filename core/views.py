#from django.shortcuts import render
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from core import models, forms
from django.contrib.staticfiles.templatetags.staticfiles import static
from django.http import HttpResponseRedirect, HttpResponse

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Protection, Font
from openpyxl.cell import get_column_letter

def create_population_data_spreadsheet():
    # Create the workbook
    wb = Workbook()

    # Create the template spreadsheet
    ws = wb.active
    ws.sheet_properties.tabColor = "1072BA"

    # Store formats
    heading = Font(bold=True)
    protected = Protection(locked=True, hidden=False)

    # Add the column + validation
    ws['A1'] = 'genus'
    ws['B1'] = 'species'
    ws['C1'] = 'count'
    ws['D1'] = 'collision_risk'
    ws['E1'] = 'density_km'
    ws['F1'] = 'passage_rate'
    ws.row_dimensions[1].font = heading
    ws.row_dimensions[1].protection = protected

    # Get a list of the valid species & genera for validation
    genera = sorted(list(models.Taxa.objects.values_list('genus', flat=True).distinct()))
    species = sorted(list(models.Taxa.objects.values_list('species', flat=True).distinct()))

    # Create additional sheets to hold them
    genus_validation_sheet = wb.create_sheet()
    genus_validation_sheet.title = 'Valid Genera'
    species_validation_sheet = wb.create_sheet()
    species_validation_sheet.title = 'Valid Species'

    # Population the cells
    for i, g in enumerate(genera, 1):
        genus_validation_sheet.cell(row=i, column=1, value=g)
    for i, s in enumerate(species, 1):
        species_validation_sheet.cell(row=i, column=1, value=s)


    # Lock the validation sheets so they cannot be tampered with
    species_validation_sheet.protection.enable()
    genus_validation_sheet.protection.enable()

    # Create & add the validation
    genera_dv = DataValidation(
        type='list',
        formula1="='Valid Genera'!A1:A" + str(len(genera)),
        error='Invalid genus. Please see the "Valid Genera" sheet to view allowed genera.',
        promptTitle='Restricted list',
        prompt='Please see the "Valid Genera" sheet to view allowed genera.'
    )
    species_dv = DataValidation(
        type='list',
        formula1="='Valid Species'!A1:A" + str(len(species)),
        error='Invalid genus. Please see the "Valid Species" sheet to view allowed species.',
        promptTitle='Restricted list',
        prompt='Please see the "Valid Species" sheet to view allowed species.'
    )
    ws.add_data_validation(genera_dv)
    ws.add_data_validation(species_dv)
    genera_dv.ranges.append('A2:A1048576')
    species_dv.ranges.append('B2:B1048576')

    # Add the list validation
    collision_risk_dv = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        promptTitle='Restricted list',
        prompt='Please enter either: "High", "Medium" or "Low".',
        error='Invalid collision risk. Please enter either: "High", "Medium" or "Low".'
    )
    ws.add_data_validation(collision_risk_dv)
    collision_risk_dv.ranges.append('D2:D1048576')

    # Add the number validation
    count_dv = DataValidation(
        type='whole',
        operator='greaterThan',
        formula1=0,
        prompt='Please enter a whole number greater than 0.',
        error='Invalid. Please enter a whole number greater than 0.'
    )
    ws.add_data_validation(count_dv)
    count_dv.ranges.append('C2:C1048576')
    density_km_dv = DataValidation(  # operator="between", formula1=0, formula2=1
        type='decimal',
        operator='greaterThan',
        formula1=0.01,
        prompt='Please enter a number greater than 0.',
        error='Invalid. Please enter a number greater than 0.'
    )
    ws.add_data_validation(density_km_dv)
    density_km_dv.ranges.append('E1:E1048576')
    passage_rate_dv = DataValidation(
        type='whole',
        operator='greaterThan',
        formula1=0,
        prompt='Please enter a whole number greater than 0.',
        error='Invalid. Please enter a whole number greater than 0.'
    )
    ws.add_data_validation(passage_rate_dv)
    passage_rate_dv.ranges.append('F2:F1048576')

    # Set column widths
    column_widths = []
    for row in ws:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell.value) > column_widths[i]:
                    column_widths[i] = len(cell.value) + 10
            else:
                column_widths += [len(cell.value) + 10]

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width

    # Close & save
    wb.save('core' + static('population_data_for_upload.xlsx'))


def create_population_data_spreadsheet_old():
    # TODO when you have time see if you can make this work with openpyxl? http://openpyxl.readthedocs.org/en/latest/validation.html
    # Create the workbook
    workbook = xlsxwriter.Workbook('core' + static('population_data_for_upload.xlsx'))

    # Create the template spreadsheet
    worksheet = workbook.add_worksheet('Main')

    # Store formats
    heading = workbook.add_format({'bold': True, 'locked': True})

    # Add the column + validation
    worksheet.write('A1', 'genus', heading)
    worksheet.write('B1', 'species', heading)
    worksheet.write('C1', 'count', heading)
    worksheet.write('D1', 'collision_risk', heading)
    worksheet.write('E1', 'density_km', heading)
    worksheet.write('F1', 'passage_rate', heading)

    # Add the list of genus + species options to an additional worksheet so that the validation works
    genus_validation_sheet = workbook.add_worksheet('Valid Genera')
    genera = sorted(list(models.Taxa.objects.values_list('genus', flat=True).distinct()))
    i = 0
    for g in genera:
        genus_validation_sheet.write(i, 0, g)
        i += 1

    species_validation_sheet = workbook.add_worksheet('Valid Species')
    species = sorted(list(models.Taxa.objects.values_list('species', flat=True).distinct()))
    j = 0
    for s in species:
        species_validation_sheet.write(j, 0, s)
        j += 1

    # Lock the validation sheet so it cannot be tampered with
    species_validation_sheet.protect()
    genus_validation_sheet.protect()

    # Add the list validation
    worksheet.data_validation('A2:A800', {
        'validate': 'list',
        'source': "='Valid Genera'!A1:A" + str(i + 1),
        'input_title': 'Restricted list',
        'input_message': 'Please see the "Valid Genera" sheet to view allowed genera.',
        'error_message': 'Invalid genus. Please see the "Valid Genera" sheet to view allowed genera.'
    })
    worksheet.data_validation('B2:B800', {
        'validate': 'list',
        'source': "='Valid Species'!A1:A" + str(j + 1),
        'input_title': 'Restricted list',
        'input_message': 'Please see the "Valid Species" sheet to view allowed species.',
        'error_message': 'Invalid species. Please see the "Valid Species" sheet to view allowed species.'
    })
    worksheet.data_validation('D2:D100', {
        'validate': 'list',
        'source': ['High', 'Medium', 'Low'],
        'input_title': 'Restricted list',
        'input_message': 'Please enter either: "High", "Medium" or "Low".',
        'error_message': 'Invalid collision risk. Please enter either: "High", "Medium" or "Low".'
    })

    # Add the number validation
    worksheet.data_validation('C2:C100', {
        'validate': 'integer',
        'criteria': '>=',
        'value': 1,
        'input_message': 'Please enter a whole number greater than 0.',
        'error_message': 'Invalid. Please enter a whole number greater than 0.'
    })
    worksheet.data_validation('E2:E100', {
        'validate': 'decimal',
        'criteria': '>=',
        'value': 0.01,
        'input_message': 'Please enter a number greater than 0.',
        'error_message': 'Invalid. Please enter a number greater than 0.'
    })
    worksheet.data_validation('F2:F100', {
        'validate': 'decimal',
        'criteria': '>',
        'value': 0.01,
        'input_message': 'Please enter a number greater than 0.',
        'error_message': 'Invalid. Please enter a number greater than 0.'
    })

    # Set column widths
    worksheet.set_column('A:F', 20)

    # Close & save
    workbook.close()


class FocalSiteDataCreate(CreateView):
    model = models.FocalSiteData
    template_name_suffix = '_create_form'
    form_class = forms.FocalSiteDataCreateForm


class FocalSiteCreate(CreateView):
    model = models.FocalSite
    template_name_suffix = '_create_form'
    form_class = forms.FocalSiteCreateForm


class ProjectList(ListView):
    model = models.Project
    context_object_name = 'projects'


class ProjectDetail(DetailView):
    model = models.Project
    context_object_name = 'project'


class ProjectCreate(CreateView):
    model = models.Project
    template_name_suffix = '_create_form'
    form_class = forms.ProjectCreateForm


class ProjectUpdate(UpdateView):
    model = models.Project
    template_name_suffix = '_update_form'
    form_class = forms.ProjectUpdateForm


class ProjectDelete(DeleteView):
    model = models.Project
    template_name_suffix = '_delete_form'
    form_class = forms.ProjectDeleteForm


class DeveloperCreate(CreateView):
    model = models.Developer
    fields = ['name', 'email', 'phone']
    template_name_suffix = '_create_form'


class DataList(ListView):
    model = models.Project
    context_object_name = 'projects'


class PopulationDataCreate(CreateView):
    model = models.PopulationData
    template_name_suffix = '_create_form'
    form_class = forms.PopulationDataCreateForm


class PopulationDataCreateView(FormView):
    template_name = 'core/populationdata_create_form.html'
    form_class = forms.MetaDataCreateForm

    create_population_data_spreadsheet()

    def form_valid(self, form):
        # This method is called when valid form data has been POSTed.
        # It should return an HttpResponse.
        print('calling form_valid')
        form.process_data(project_pk=self.kwargs['project_pk'])
        #return super(PopulationDataCreateView, self).form_valid(form)
        return HttpResponseRedirect(self.get_success_url())



'''def create_population_data(request, project_pk):
    metadata_form = forms.MetaDataCreateForm()

    if request.POST:
            metadata_form = forms.MetaDataCreateForm(request.POST, prefix='meta')
            data_form = forms.PopulationDataCreateForm(request.POST, request.FILES, prefix='meta')

            if metadata_form.is_valid() and data_form.is_valid():
                    # Prepare & save the metadata
                    metadata = metadata_form.save(commit=False)
                    metadata.project = models.Project.objects.get(pk=project_pk)
                    metadata.save()

                    # Parse the spreadsheet and get the data
                    data = request.FILES['spreadsheet']

                    # Prepare & save the data
                    data = data_form.save(commit=False)
                    data.metadata = metadata
                    data.save()

    # Create a spreadsheet with validation for them on-the-fly - note we may need to change this if it impacts performance
    workbook = xlsxwriter.Workbook('population_data_for_upload.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.write('A1', 'Hello world')
    workbook.close()

    return render_to_response('core/populationdata_create_form.html', {
        'metadata_form': metadata_form,
        'data_form': data_form,
        'project_pk': project_pk,
        },
        RequestContext(request))'''

