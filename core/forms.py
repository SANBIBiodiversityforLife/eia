from django.utils.safestring import mark_safe
from django import forms
from django.db import models
#from django.core.exceptions import DoesNotExist
from leaflet.forms.widgets import LeafletWidget
from core import models
from core import validators
from openpyxl import load_workbook
#from django.contrib.staticfiles.templatetags.staticfiles import static
from openpyxl.styles import Color, Fill, Style, fills, PatternFill, Font
from django.core.exceptions import ValidationError
from django.contrib.staticfiles.templatetags.staticfiles import static
import tempfile
import os
import posixpath
from django.conf import settings
from core.spreadsheet_creation import add_focal_site_data_validation, add_population_data_validation, add_fatality_data_validation
from django.contrib.auth import get_user_model
import time
from django.contrib.gis.geos import Point


class SignupForm(forms.ModelForm):
    first_name = forms.CharField(max_length=100)
    last_name = forms.CharField(max_length=100)

    class Meta:
        model = models.Profile
        fields = ('first_name', 'last_name', 'phone', 'type')

    # A custom method required to work with django-allauth, see https://stackoverflow.com/questions/12303478/how-to-customize-user-profile-when-using-django-allauth
    def signup(self, request, user):
        # Save your user
        user.first_name = self.cleaned_data['first_name']
        user.last_name = self.cleaned_data['last_name']
        user.save()

        # Save your profile
        profile = models.Profile()
        profile.user = user
        profile.phone = self.cleaned_data['phone']
        profile.type = self.cleaned_data['type']
        profile.save()


class ProfileUpdateForm(forms.ModelForm):
    email = forms.EmailField()
    first_name = forms.CharField(max_length=100)
    last_name = forms.CharField(max_length=100)

    class Meta:
        model = models.Profile
        fields = ('first_name', 'last_name', 'email', 'phone', 'type')

    def __init__(self, *args, **kwargs):
        super(ProfileUpdateForm, self).__init__(*args, **kwargs)
        try:
            self.fields['email'].initial = self.instance.user.email
            self.fields['first_name'].initial = self.instance.user.first_name
            self.fields['last_name'].initial = self.instance.user.last_name
        except models.User.DoesNotExist:
            pass

    def save(self, *args, **kwargs):
        """
        Update the primary email address on the related User object as well.
        """
        u = self.instance.user
        u.email = self.cleaned_data['email']
        u.save()
        profile = super(ProfileUpdateForm, self).save(*args,**kwargs)
        return profile


class ProjectCreateForm(forms.ModelForm):
    class Meta:
        model = models.Project
        fields = ('name', 'location', 'developer', 'eia_number', 'energy_type')
        widgets = {'location': LeafletWidget()}


class DeveloperCreateForm(forms.ModelForm):
    class Meta:
        model = models.Developer
        fields = ('name', 'email', 'phone')


class EquipmentMakeCreateForm(forms.ModelForm):
    class Meta:
        model = models.EquipmentMake
        fields = ('name',)


class FocalSiteCreateForm(forms.ModelForm):
    class Meta:
        model = models.FocalSite
        fields = ('location', 'name', 'sensitive', 'activity', 'habitat') # todo do i need to insert 'taxon' again?
        widgets = {'location': LeafletWidget()}

    def __init__(self, *args, **kwargs):
        self.project_pk = kwargs.pop('project_pk')
        self.uploader = kwargs.pop('uploader')
        super(FocalSiteCreateForm, self).__init__(*args, **kwargs)

    def clean(self):
        # Create the metadata object and store it to get its primary key
        # This must get deleted after this function if no actual data is stored
        print("adding project to instance...")
        cleaned_data = super(FocalSiteCreateForm, self).clean()
        self.instance.project = models.Project.objects.get(pk=self.project_pk)


class ProjectUpdateOperationalInfoForm(forms.ModelForm):
    class Meta:
        model = models.Project
        fields = ('operational_date', 'construction_date', 'turbine_locations', 'equipment_make', 'equipment_capacity',
                  'equipment_height')
        widgets = {'turbine_locations': LeafletWidget()}
        help_texts = {
            'operational_date': 'The date the project first became operational',
            'construction_date': 'The date construction began on the project',
            'turbine_locations': 'The locations of the turbines',
            'equipment_make': 'The make of the turbines/solar panels',
            'equipment_capacity': 'The capacity of the turbines/solar panels',
            'equipment_height': 'The height of the turbines/solar panels',
        }


class ProjectUpdateForm(forms.ModelForm):
    class Meta:
        model = models.Project
        fields = ('name', 'location', 'developer', 'eia_number', 'energy_type')
        widgets = {'location': LeafletWidget()}


class ProjectDeleteForm(forms.ModelForm):
    class Meta:
        model = models.Project
        fields = ('name', 'location', 'developer', 'eia_number', 'energy_type')
        widgets = {'location': LeafletWidget()}


class DataUploadForm(forms.Form):
    spreadsheet = forms.FileField()


class DataViewForm(forms.Form):
    datasets = forms.ChoiceField(label='Choose a dataset to view', choices=())

    def __init__(self, *args, **kwargs):
        metadata = kwargs.pop('metadata')
        super(DataViewForm, self).__init__(*args, **kwargs)
        self.fields['datasets'].choices = [(d.id, d) for d in metadata]


class RemovalFlagCreateForm(forms.ModelForm):
    class Meta:
        model = models.RemovalFlag
        fields = ('reason', 'requested_by', 'metadata')
        widgets = {'requested_by': forms.HiddenInput(), 'metadata': forms.HiddenInput()}


def write_error(main_sheet, row_number, error_message):
    main_sheet.cell(column=6, row=row_number, value=error_message)
    main_sheet.cell(column=6, row=row_number).style = Style(font=Font(color='FFFFFFFF'),
                                                            fill=PatternFill(patternType='solid', fgColor=Color('FFFF0000')))


class MetaDataCreateForm(forms.ModelForm):
    upload_data = forms.FileField(
        label=mark_safe('Upload spreadsheet'),
        validators=[validators.validate_spreadsheet]
    )
    number_of_cols = 10

    def __init__(self, *args, **kwargs):
        self.project_pk = kwargs.pop('project_pk')
        self.uploader = kwargs.pop('uploader')
        super(MetaDataCreateForm, self).__init__(*args, **kwargs)

    class Meta:
        model = models.MetaData
        fields = ('upload_data', 'control_data')

    # Gets overwritten
    def create_data_object(self, metadata, taxa, cells):
        raise ValidationError('Incorrectly calling the parent class')

    # Gets overwritten
    def add_data_validation(self, wb):
        raise ValidationError('Incorrectly calling the parent class')

    # The main function which saves all of the different data objects
    def process_data(self):
        start = time.clock()
        # Load the workbook from the file held in memory
        uploaded_data = load_workbook(self.files['upload_data'])
        print('loaded workbook - ' + str(time.clock() - start))

        # Get the correct sheet - TODO how are we going to stop them from renaming the sheet?
        main_sheet = uploaded_data.get_sheet_by_name("Main")

        # Create the metadata object and store it to get its primary key
        # This must get deleted after this function if no actual data is stored
        self.instance.project = models.Project.objects.get(pk=self.project_pk)
        self.instance.uploader = self.uploader
        self.instance.save()
        metadata = self.instance

        # Keep track of the errors somehow
        row_with_error_count = 0

        # Keep track of all the population data objects - if we have no errors at the end we shall save them
        data_object_list = []

        # Loop through the rows in the sheet
        for row in main_sheet.iter_rows(row_offset=1):
            # If all of these are blank, (i.e., none have a value), then ignore this row
            cells = [x.value for x in row[0:self.number_of_cols]]
            if not(any(cells)):
                continue
            print(self.number_of_cols)
            # If any of these are blank (i.e. any don't have a value), throw up an error and get them to fill it in
            if not(all(cells)):
                write_error(main_sheet=main_sheet,
                            row_number=row[0].row,
                            error_message='Incomplete row. All the fields must be filled out.')
                row_with_error_count += 1
                continue

            # Try and retrieve the taxa based on species name
            try:
                # taxa = models.Taxon.objects.get(genus__iexact=cells[0], species__iexact=cells[1])
                taxa = models.Taxon.objects.get(name__iexact=cells[0])
            except models.Taxon.DoesNotExist:
                write_error(main_sheet=main_sheet,
                            row_number=row[0].row,
                            error_message='Error with species - does not exist. Please check and correct.')
                row_with_error_count += 1
                continue

            # Try and save the data object (e.g. population data or whatever it is)
            try:
                # Create a new object so we can run the validation on it
                data_object = self.create_data_object(metadata, taxa, cells)

                # Call the validation on the object
                data_object.full_clean()

                # If it hasn't slipped into the except, add it to the main list
                data_object_list.append(data_object)

                # Remove any validation left over from a previous upload
                main_sheet.cell(column=6, row=row[0].row, value='')
                main_sheet.cell(column=6, row=row[0].row).style = Style(font=Font(color='00000000'),
                                                                        fill=PatternFill(patternType='solid', fgColor=Color('FFFFFF00')))

            except ValidationError as err:
                # Write an error
                write_error(main_sheet=main_sheet,
                            row_number=row[0].row,
                            error_message='Error when saving {}'.format(err))
                row_with_error_count += 1
                continue

        if row_with_error_count > 0:
            print('Errors with data - extra processing')
            # Delete the metadata
            metadata.delete()

            # Add the validation
            print('Adding validation')
            self.add_data_validation(uploaded_data)
            print('/Adding validation')

            # Unique filename and timestamp - used os.path.join but need linux paths for absolute urls
            tmp_dir = posixpath.join(settings.BASE_DIR, 'core', 'static', 'core', 'tmp')
            fd, unique_file = tempfile.mkstemp(suffix='.xlsx', prefix='data_', dir=tmp_dir)

            # TODO serve the files through a proper webserver
            uploaded_data.save(unique_file)
            os.close(fd)
            print('Created error file, returning URL')

            # Send back the absolute file url
            return posixpath.join(posixpath.sep, 'static', 'core', 'tmp', os.path.basename(os.path.relpath(unique_file)))
        else:
            print('No errors - saving data')
            # Save all the validated/cleaned data
            for data_object in data_object_list:
                data_object.save()

            # No errors, return
            return False


class PopulationDataCreateForm(MetaDataCreateForm):
    number_of_cols = 5

    def create_data_object(self, metadata, taxa, cells):
        return models.PopulationData(metadata=metadata,
                              taxa=taxa,
                              count=cells[1],
                              collision_risk=cells[2],
                              density_km=cells[3],
                              passage_rate=cells[4])

    def add_data_validation(self, wb):
        add_population_data_validation(wb)


class FocalSiteDataCreateForm(MetaDataCreateForm):
    number_of_cols = 4

    def __init__(self, *args, **kwargs):
        self.focal_site_pk = kwargs.pop('focal_site_pk')
        super(FocalSiteDataCreateForm, self).__init__(*args, **kwargs)

    def create_data_object(self, metadata, taxa, cells):
        return models.FocalSiteData(metadata=metadata,
                             taxa=taxa,
                             count=cells[1],
                             life_stage=cells[2],
                             activity=cells[3],
                             focal_site=models.FocalSite.objects.get(pk=self.focal_site_pk))

    def add_data_validation(self, wb):
        add_focal_site_data_validation(wb)


class FatalityDataCreateForm(MetaDataCreateForm):
    number_of_cols = 4

    def create_data_object(self, metadata, taxa, cells):
        return models.FatalityData(metadata=metadata,
                            taxa=taxa,
                            coordinates=Point(cells[2], cells[1]),
                            cause_of_death=cells[3])

    def add_data_validation(self, wb):
        add_fatality_data_validation(wb)
