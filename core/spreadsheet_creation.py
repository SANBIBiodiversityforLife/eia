from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Protection, Font, Style

from openpyxl.cell import get_column_letter
from core import models
from django.contrib.staticfiles.templatetags.staticfiles import static
from django.conf import settings

import requests


def reset_taxa_tree(request):
    # Reset spreadsheets (move this down later)
    create_population_data_spreadsheet()
    create_fatality_data_spreadsheet()
    create_focal_site_data_spreadsheet()
    create_population_data_spreadsheet(validation=False)
    create_fatality_data_spreadsheet(validation=False)
    create_focal_site_data_spreadsheet(validation=False)

    # Just temporarily making sure I don't delete it by accident
    return

    # Empty taxon table of everything apart from root items
    models.Taxon.objects.filter(is_root=False).delete()

    # Select everything in the DB but exclude Chiroptera order (bats) & Aves class (birds)
    taxa = models.Taxon.objects.exclude(name__in=settings.BASE_TAXA)

    # Now all we have left is the items which we need to rebuild the db from
    for taxon in taxa:
        recursive_tree_builder(taxon.id)

    # Reset spreadsheets

def recursive_tree_builder(parent_id):
    offset = 0
    # r = requests.get(settings.GBIF_API_URL.format(id=parent_id, offset=offset))
    # data = r.json()
    end_of_records = False

    while not end_of_records:
        # Get the data
        r = requests.get(settings.GBIF_API_URL.format(id=parent_id, offset=offset))
        data = r.json()

        # First, loop through the children in this result set
        children = data['results']
        for child in children:
            taxon_id = int(child['nubKey'])

            # Get the rank key
            rank = False
            for key, value in dict(models.Taxon.RANK_CHOICES).items():
                if value.lower() == child['rank'].lower():
                    rank = key

            # If we couldn't find it in our list throw an exception
            if not rank:
                raise ValueError('Rank does not exist in database: ' + child['rank'] + ' For taxon ' +
                                 child['canonicalName'] + '(' + child['rank'] + ') - ' + str(taxon_id))

            print(child['canonicalName'] + '(' + child['rank'] + ') - ' + str(taxon_id))

            # If it's a family, species, genus then check and see if there are any occurrence
            # records in SA before proceeding. If there are none then go onto the next item in the loop.
            # Don't want to check subsp, variety etc in case they are unknown/rare/new and have no occurrence records
            if rank in [models.Taxon.FAMILY, models.Taxon.GENUS, models.Taxon.SPECIES]:
                r = requests.get(settings.GBIF_API_OCCURRENCE_URL.format(id=taxon_id))
                occurrences = r.json()
                if occurrences['count'] == 0:
                    print(' Not in SA : ' + child['canonicalName'] + '(' + child['rank'] + ') - ' + str(taxon_id))
                    continue
                else:
                    print(' --- IN SA : ' + child['canonicalName'] + '(' + child['rank'] + ') - ' + str(taxon_id))

            # Populate and save the taxon object
            taxon = models.Taxon(name=child['canonicalName'],
                                 rank=rank,
                                 id=taxon_id,
                                 parent_id=parent_id)
            if 'vernacularName' in child:
                taxon.vernacular_name = child['vernacularName']
            taxon.save()

            # Recursion
            recursive_tree_builder(taxon_id)

        # Are we at the end of the records? if so we need to break out of the loop
        end_of_records = data['endOfRecords']

        # Increase the offset in case we are not at the end of records
        offset += settings.GBIF_API_OFFSET



def add_taxa_validation(wb):
    ws = wb.get_sheet_by_name("Main")
    species_validation_sheet = wb.get_sheet_by_name("Valid Species")

    # Add the protect to the sheet, and remove it for the individual cells
    ws.protection.sheet = True
    for row in ws.iter_rows('A2:F' + str(settings.MAX_XLSX_ROWS)):
        for cell in row:
            cell.style = Style(protection=Protection(locked=False, hidden=False))

    # Lock the validation sheets so they cannot be tampered with
    species_validation_sheet.protection.enable()

    # Species validation
    species_dv = DataValidation(
        type='list',
        formula1="='Valid Species'!A1:A" + str(species_validation_sheet.max_row),
        error='Invalid genus. Please see the "Valid Species" sheet to view allowed species.',
        promptTitle='Restricted list',
        prompt='Please see the "Valid Species" sheet to view allowed species.'
    )
    ws.add_data_validation(species_dv)
    species_dv.ranges.append('A2:A1048576')


def create_template_spreadsheet(validation):
    # Create the workbook
    wb = Workbook()

    # Create the template spreadsheet
    ws = wb.active
    ws.title = 'Main'
    ws.sheet_properties.tabColor = "1072BA"

    # Freeze panes
    ws.freeze_panes = ws['A2']

    # Add the columns
    ws['A1'] = 'name'

    # Format them in bold
    heading = Style(font=Font(bold=True), protection=Protection(locked=True, hidden=False))
    for row in ws.iter_rows('A1:F1'):
        for cell in row:
            cell.style = heading

    # Get a list of the valid species for validation
    species_ranks = [models.Taxon.SPECIES, models.Taxon.INFRASPECIFIC_NAME, models.Taxon.SUBSPECIES]
    species = sorted(list(models.Taxon.objects.filter(rank__in=species_ranks).values_list('name', flat=True).distinct()))
    # species = sorted(list(models.Taxa.objects.values_list('species', flat=True).distinct()))

    # Create additional sheets to hold them
    species_validation_sheet = wb.create_sheet()
    species_validation_sheet.title = 'Valid Species'

    # Population the cells
    for i, s in enumerate(species, 1):
        species_validation_sheet.cell(row=i, column=1, value=s)

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 100

    return wb


def add_count_validation(ws):
    # Count is used for pop and focal site data, so let's separate it out here
    count_dv = DataValidation(
        type='whole',
        operator='greaterThan',
        formula1=0,
        prompt='Please enter a whole number greater than 0.',
        error='Invalid. Please enter a whole number greater than 0.'
    )
    ws.add_data_validation(count_dv)
    count_dv.ranges.append('B2:B' + str(settings.MAX_XLSX_ROWS))


def format_list(list_for_formatting):
    if len(list_for_formatting) > 1:
        return ', '.join(list_for_formatting[:-1]) + ' or ' + list_for_formatting[-1]
    elif len(list_for_formatting) == 1:
        return list_for_formatting[0]
    else:
        return 'Error'


def list_validation(tuple_list):
    # We are expecting something like [('D', 'Description),]
    formula = [x[0] for x in tuple_list]
    verbose = [x[1] for x in tuple_list]

    # Format it into a nice prompt text
    prompt_text = 'Please enter either: ' + format_list(formula) + ' (' + format_list(verbose) + ').'

    # Excel breaks if you add error messages longer than a certain number of chars it seems (256 maybe?)
    if len(prompt_text) > 240:
        prompt_text = prompt_text[:237] + '...'

    # Return the object
    return DataValidation(
        type="list",
        formula1='"' + ','.join(formula) + '"',
        promptTitle='Restricted list',
        prompt=prompt_text,
        error='Invalid value. ' + prompt_text
    )


def add_focal_site_data_validation(wb):
    ws = wb.active

    # Taxa
    add_taxa_validation(wb)

    # Count
    add_count_validation(ws)

    # Life stage
    dv = list_validation(models.FocalSiteData.life_stage_choices)
    ws.add_data_validation(dv)
    dv.ranges.append('C2:C' + str(settings.MAX_XLSX_ROWS))

    # Activity
    dv = list_validation(models.FocalSiteData.activity_choices)
    ws.add_data_validation(dv)
    dv.ranges.append('D2:D' + str(settings.MAX_XLSX_ROWS))


def add_fatality_data_validation(wb):
    ws = wb.active

    # Taxa
    add_taxa_validation(wb)

    # Latitude
    latitude_dv = DataValidation(
        type='decimal',
        operator='between',
        formula1=-35,
        formula2=-21,
        prompt='Please enter a negative number between -21 and -35.',
        error='Invalid. Please enter a negative number between -21 and -35.'
    )
    ws.add_data_validation(latitude_dv)
    latitude_dv.ranges.append('B1:B' + str(settings.MAX_XLSX_ROWS))

    # Longitude
    longitude_dv = DataValidation(
        type='decimal',
        operator='between',
        formula1=16,
        formula2=33,
        prompt='Please enter a number between 16 and 31.',
        error='Invalid. Please enter a number between 16 and 31.'
    )
    ws.add_data_validation(longitude_dv)
    longitude_dv.ranges.append('C1:C' + str(settings.MAX_XLSX_ROWS))

    # Cause of death
    dv = list_validation(models.FatalityData.cause_of_death_choices)
    ws.add_data_validation(dv)
    dv.ranges.append('D2:D' + str(settings.MAX_XLSX_ROWS))


def add_population_data_validation(wb):
    ws = wb.get_sheet_by_name("Main")

    # Taxa
    add_taxa_validation(wb)

    # Count
    add_count_validation(ws)

    # Collision risk
    dv = list_validation(models.PopulationData.collision_risk_choices)
    ws.add_data_validation(dv)
    dv.ranges.append('C2:C' + str(settings.MAX_XLSX_ROWS))

    # Density
    density_km_dv = DataValidation(  # operator="between", formula1=0, formula2=1
        type='decimal',
        operator='greaterThan',
        formula1=0.01,
        prompt='Please enter a number greater than 0.',
        error='Invalid. Please enter a number greater than 0.'
    )
    ws.add_data_validation(density_km_dv)
    density_km_dv.ranges.append('D1:D' + str(settings.MAX_XLSX_ROWS))

    # Passage rate
    passage_rate_dv = DataValidation(
        type='whole',
        operator='greaterThan',
        formula1=0,
        prompt='Please enter a whole number greater than 0.',
        error='Invalid. Please enter a whole number greater than 0.'
    )
    ws.add_data_validation(passage_rate_dv)
    passage_rate_dv.ranges.append('E2:E' + str(settings.MAX_XLSX_ROWS))


def create_population_data_spreadsheet(validation=True):
    wb = create_template_spreadsheet(validation)
    ws = wb.active

    # Add the additional fields, A and B are always species + genus
    ws['B1'] = 'count'
    ws['C1'] = 'collision_risk'
    ws['D1'] = 'density_km'
    ws['E1'] = 'passage_rate'

    # Add the validation
    if validation:
        add_population_data_validation(wb)

        # Close & save
        wb.save('core' + static('population_data_for_upload.xlsx'))
    else:
        wb.save('core' + static('population_data_no_validation.xlsx'))


def create_focal_site_data_spreadsheet(validation=True):
    wb = create_template_spreadsheet(validation)
    ws = wb.active

    # Add the additional fields, A and B are always species + genus
    ws['B1'] = 'count'
    ws['C1'] = 'life_stage'
    ws['D1'] = 'activity'

    # Add the validation
    if validation:
        add_focal_site_data_validation(wb)

        # Close & save
        wb.save('core' + static('focal_site_data_for_upload.xlsx'))
    else:
        wb.save('core' + static('focal_site_data_no_validation.xlsx'))


def create_fatality_data_spreadsheet(validation=True):
    wb = create_template_spreadsheet(validation)
    ws = wb.active

    # Add the additional fields, A and B are always species + genus
    ws['B1'] = 'latitude'
    ws['C1'] = 'longitude'
    ws['D1'] = 'cause_of_death'

    # Add the validation
    if validation:
        add_fatality_data_validation(wb)

        # Close & save
        wb.save('core' + static('fatality_data_for_upload.xlsx'))
    else:
        wb.save('core' + static('fatality_data_no_validation.xlsx'))
